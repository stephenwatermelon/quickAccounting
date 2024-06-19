import openpyxl
import json
import re

def read_properties(file_path):
    properties = {}
    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                key, value = line.split('=', 1)
                properties[key.strip()] = int(value.strip())
    return properties

def clean_data(value):
    if isinstance(value, str):
        # Remove full-width parentheses and any content within them
        value = re.sub(r'\uff08.*?\uff09', '', value)
    elif isinstance(value, float):
        # Round float to two decimal places
        value = round(value, 2)
    return value

def is_blank_row(row, properties):
    for key in properties.values():
        if key < len(row) and row[key] is not None:
            return False
    return True

def read_insurance_excel(file_path, sheet_names, properties, start_row, end_row):
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = []

        # Validate start and end rows
        if start_row < 1:
            raise ValueError("start_row must be greater than 0")
        if end_row > sheet.max_row:
            raise ValueError(f"end_row must be less than or equal to {sheet.max_row}")
        if start_row > end_row:
            raise ValueError("start_row must be less than or equal to end_row")

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            if is_blank_row(row, properties):
                continue
            try:
                row_data = {
                    'level': clean_data(row[properties['level']]) if properties['level'] < len(row) else None,
                    'minAmount': clean_data(row[properties['minAmount']]) if properties['minAmount'] < len(row) else None,
                    'maxAmount': clean_data(row[properties['maxAmount']]) if properties['maxAmount'] < len(row) else None,
                    'healthInsuranceFeesUnder40': clean_data(row[properties['healthInsuranceFeesUnder40']]) if properties['healthInsuranceFeesUnder40'] < len(row) else None,
                    'healthInsuranceFeesOver40': clean_data(row[properties['healthInsuranceFeesOver40']]) if properties['healthInsuranceFeesOver40'] < len(row) else None,
                    'pensionInsuranceFees': clean_data(row[properties['pensionInsuranceFees']]) if properties['pensionInsuranceFees'] < len(row) else None,
                }
                sheet_data.append(row_data)
            except KeyError as e:
                print(f"KeyError: {e} not found in properties.")
            except IndexError as e:
                print(f"IndexError: {e}. Row: {row}")

        data[sheet_name] = sheet_data
    
    return data
    
def read_tax_excel(file_path, sheet_names, properties, start_row, end_row):
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = []

        # Validate start and end rows
        if start_row < 1:
            raise ValueError("start_row must be greater than 0")
        if end_row > sheet.max_row:
            raise ValueError(f"end_row must be less than or equal to {sheet.max_row}")
        if start_row > end_row:
            raise ValueError("start_row must be less than or equal to end_row")

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            if is_blank_row(row, properties):
                continue
            try:
                row_data = {
                    'level': clean_data(row[properties['level']]) if properties['level'] < len(row) else None,
                    'minAmount': clean_data(row[properties['minAmount']]) if properties['minAmount'] < len(row) else None,
                    'maxAmount': clean_data(row[properties['maxAmount']]) if properties['maxAmount'] < len(row) else None,
                    'dependents0': clean_data(row[properties['dependents0']]) if properties['dependents0'] < len(row) else None,
                    'dependents1': clean_data(row[properties['dependents1']]) if properties['dependents1'] < len(row) else None,
                    'dependents2': clean_data(row[properties['dependents2']]) if properties['dependents2'] < len(row) else None,
                    'dependents3': clean_data(row[properties['dependents3']]) if properties['dependents3'] < len(row) else None,
                    'dependents4': clean_data(row[properties['dependents4']]) if properties['dependents4'] < len(row) else None,
                    'dependents5': clean_data(row[properties['dependents5']]) if properties['dependents5'] < len(row) else None,
                    'dependents6': clean_data(row[properties['dependents6']]) if properties['dependents6'] < len(row) else None,
                    'dependents7': clean_data(row[properties['dependents7']]) if properties['dependents7'] < len(row) else None,
                    
                }
                sheet_data.append(row_data)
            except KeyError as e:
                print(f"KeyError: {e} not found in properties.")
            except IndexError as e:
                print(f"IndexError: {e}. Row: {row}")
    
    return sheet_data

def write_json(data, output_file):
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=4)


def generate_insurance_06():
    insurance_properties_file = 'E:/quickAccounting/src/main/resources/initData/insurance_properties.txt'
    insurance_excel_file = 'E:/quickAccounting/src/main/resources/initData/insurance06.xlsx'
    insurance_json_file = 'E:/quickAccounting/src/main/resources/initData/insuranceData06.json'

    # List of sheet names to read
    sheet_names = ['東京', '千葉', '神奈川', '埼玉', '青森', '秋田', '大阪']  # Add your sheet names here

    # Designate start and end rows (inclusive)
    start_row = 12  # assuming you want to start from the second row (first row is header)
    end_row = 61  # change this to the desired end row number

    properties = read_properties(insurance_properties_file)

    insurance_data = read_insurance_excel(insurance_excel_file, sheet_names, properties, start_row, end_row)

    write_json(insurance_data, insurance_json_file)

def generate_insurance_05():
    insurance_properties_file = 'E:/quickAccounting/src/main/resources/initData/insurance_properties.txt'
    insurance_excel_file = 'E:/quickAccounting/src/main/resources/initData/insurance05.xlsx'
    insurance_json_file = 'E:/quickAccounting/src/main/resources/initData/insuranceData05.json'

    # List of sheet names to read
    sheet_names = ['東京', '千葉', '神奈川', '埼玉', '青森', '秋田', '大阪']  # Add your sheet names here

    # Designate start and end rows (inclusive)
    start_row = 12  # assuming you want to start from the second row (first row is header)
    end_row = 61  # change this to the desired end row number

    properties = read_properties(insurance_properties_file)

    insurance_data = read_insurance_excel(insurance_excel_file, sheet_names, properties, start_row, end_row)

    write_json(insurance_data, insurance_json_file)
    
def generate_tax_06():
    tax_properties_file = 'E:/quickAccounting/src/main/resources/initData/tax_properties.txt'
    tax_excel_file = 'E:/quickAccounting/src/main/resources/initData/tax06.xlsx'
    tax_json_file = 'E:/quickAccounting/src/main/resources/initData/taxData06.json'

    #TAX
    properties = read_properties(tax_properties_file)
    sheet_names = ['月額表']
    start_row = 10  
    end_row = 302 
    tax_data = read_tax_excel(tax_excel_file, sheet_names, properties, start_row, end_row)
    write_json(tax_data, tax_json_file)

def generate_tax_05():
    tax_properties_file = 'E:/quickAccounting/src/main/resources/initData/tax_properties.txt'
    tax_excel_file = 'E:/quickAccounting/src/main/resources/initData/tax05.xlsx'
    tax_json_file = 'E:/quickAccounting/src/main/resources/initData/taxData05.json'

    #TAX
    properties = read_properties(tax_properties_file)
    sheet_names = ['月額表']
    start_row = 10  
    end_row = 302 
    tax_data = read_tax_excel(tax_excel_file, sheet_names, properties, start_row, end_row)
    write_json(tax_data, tax_json_file)

def main():
    generate_insurance_06()
    generate_insurance_05()
    generate_tax_06()
    generate_tax_05()
    

if __name__ == '__main__':
    main()
