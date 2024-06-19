import openpyxl
import json
import re

def read_properties(file_path):
    properties = {}
    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                key, value = line.split('=', 1)
                properties[key.strip()] = int(value.strip())
    return properties

def clean_data(value):
    if isinstance(value, str):
        # Remove full-width parentheses and any content within them
        value = re.sub(r'/uff08.*?/uff09', '', value)
        # You can add more replacements here if needed
    return value

def read_insurance_excel(file_path, sheet_names, properties, start_row, end_row):
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = []

        # Validate start and end rows
        if start_row < 1:
            raise ValueError("start_row must be greater than 0")
        if end_row > sheet.max_row:
            raise ValueError(f"end_row must be less than or equal to {sheet.max_row}")
        if start_row > end_row:
            raise ValueError("start_row must be less than or equal to end_row")

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            try:
                row_data = {
                    'level': clean_data(row[properties['level']]),
                    #'amount': row[properties['amount']],
                    'minAmount': row[properties['minAmount']],
                    'maxAmount': row[properties['maxAmount']],
                    'healthInsuranceFeesUnder40': row[properties['healthInsuranceFeesUnder40']],
                    'healthInsuranceFeesOver40': row[properties['healthInsuranceFeesOver40']],
                    'pensionInsuranceFees': row[properties['pensionInsuranceFees']],
                }
                sheet_data.append(row_data)
            except KeyError as e:
                print(f"KeyError: {e} not found in properties.")
            except IndexError as e:
                print(f"IndexError: {e}. Row: {row}")

        data[sheet_name] = sheet_data
    
    return data
    
def read_tax_excel(file_path, sheet_names, properties, start_row, end_row):
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = []

        # Validate start and end rows
        if start_row < 1:
            raise ValueError("start_row must be greater than 0")
        if end_row > sheet.max_row:
            raise ValueError(f"end_row must be less than or equal to {sheet.max_row}")
        if start_row > end_row:
            raise ValueError("start_row must be less than or equal to end_row")

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            try:
                row_data = {
                    'level': clean_data(row[properties['level']]),
                    'minAmount': row[properties['minAmount']],
                    'maxAmount': row[properties['maxAmount']],
                    'dependents0': row[properties['dependents0']],
                    'dependents1': row[properties['dependents1']],
                    'dependents2': row[properties['dependents2']],
                    'dependents3': row[properties['dependents3']],
                    'dependents4': row[properties['dependents4']],
                    'dependents5': row[properties['dependents5']],
                    'dependents6': row[properties['dependents6']],
                    'dependents7': row[properties['dependents7']],
                }
                sheet_data.append(row_data)
            except KeyError as e:
                print(f"KeyError: {e} not found in properties.")
            except IndexError as e:
                print(f"IndexError: {e}. Row: {row}")

        data[sheet_name] = sheet_data
    
    return data

def write_json(data, output_file):
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=4)
        
        

def main():
    insurance_properties_file = 'E:/quickAccounting/src/main/resources/insurance_properties.txt'
    tax_properties_file = 'E:/quickAccounting/src/main/resources/tax_properties.txt'
    insurance_excel_file = 'E:/quickAccounting/src/main/resources/insurance.xlsx'
    tax_excel_file = 'E:/quickAccounting/src/main/resources/tax.xlsx'
    insurance_json_file = 'E:/quickAccounting/src/main/resources/insuranceData.json'
    tax_json_file = 'E:/quickAccounting/src/main/resources/taxData.json'

    # List of sheet names to read
    sheet_names = ['東京', '千葉', '神奈川', '埼玉', '青森', '秋田', '大阪']  # Add your sheet names here

    # Designate start and end rows (inclusive)
    start_row = 12  # assuming you want to start from the second row (first row is header)
    end_row = 61  # change this to the desired end row number

    properties = read_properties(insurance_properties_file)
    print(f"Properties: {properties}")

    insurance_data = read_insurance_excel(insurance_excel_file, sheet_names, properties, start_row, end_row)
    print(f"Data: {insurance_data}")

    write_json(insurance_data, insurance_json_file)
    
    
    #TAX
    properties = read_properties(tax_properties_file)
    sheet_names = ['月額表']
    start_row = 10  
    end_row = 302 
    tax_data = read_tax_excel(tax_excel_file, sheet_names, properties, start_row, end_row)
    write_json(tax_data, tax_json_file)
    

if __name__ == '__main__':
    main()
